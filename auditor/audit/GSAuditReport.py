from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import re
import pandas as pd


def highlight_even_row(*, row) -> list:
    return ['background-color: gainsboro;' if _ % 2 == 1 else '' for _ in range(len(row))]
    #     border: 1px solid


def gs_audit_report_color_rows(*, row) -> list:
    if row['flag'] in ['True', 'TRUE', True]: return ['' for _ in row]
    else:
        if row['GSValue'].startswith('GS_Error'):
            return ['background-color: red; color: white; border: 2px solid green' for _ in row]
        elif row['Permission'].lower().startswith('global'):
            return ['background-color: red' for _ in row]
        elif row['Permission'].lower().startswith('local'):
            return ['background-color: cyan' for _ in row]
        elif row['Permission'].lower().startswith('not auditable'):
            return ['background-color: dimgray' for _ in row]
        else:
            return ['background-color: crimson' for _ in row]


thin_black_side = Side(border_style="thin", color="000000")
thin_border = Border(left=thin_black_side, right=thin_black_side, top=thin_black_side, bottom=thin_black_side)
tmp_sheet_dict = {
    'Defult': ['Defult'],
    'LTE': ['LTE', 'LTE_RATFreqPrio'],
    'LTECell': ['LTECell'],
    'LTERelation': ['LTERelation_EUtran', 'LTERelation_Utran', 'LTERelation_GUtran', 'LTENR_para'],
    'NR': ['NR'],
    'NRCell': ['NRCell'],
    'NRRelation': ['NRRelation'],
}


class GSAuditReport:
    def __init__(self, usid, gs_admin: bool = False):
        self.usid = usid
        df_temp = self.usid.df_report.copy(deep=True)
        if len(df_temp.index) == 0: return
        self.audit_file = pd.ExcelWriter(os.path.join(self.usid.outdir, F'{self.usid.site_id}_{self.usid.revision}.xlsx'), engine='openpyxl')
        df_temp = df_temp[['Site', 'MO', 'MOC', 'Type', 'Parameter', 'CurrentValue', 'GSValue', 'InitialValue', 'Permission', 'Suffix', 'flag']]
        df_temp.MO = df_temp.MO.str.extract(r'.*,ManagedElement=\w+,(.*)').squeeze()
        for sheet in tmp_sheet_dict:
            if 'Relation' in sheet:
                df = df_temp.loc[(df_temp.Type.str.startswith(sheet, na=False))]
            else:
                df = df_temp.loc[(df_temp.Type.isin(tmp_sheet_dict.get(sheet)))]
            if df.shape[0] > 0:
                df = df[['Site', 'MOC', 'Parameter', 'MO', 'CurrentValue', 'GSValue', 'InitialValue', 'Permission', 'Suffix', 'flag']]
                df_style = df.style.apply(lambda x: highlight_even_row(row=x))
                df_style = df_style.apply(lambda x: gs_audit_report_color_rows(row=x), axis=1)
                df_style.to_excel(excel_writer=self.audit_file, sheet_name=sheet, index=False, header=True)
                self.audit_file.sheets[sheet].auto_filter.ref = self.audit_file.sheets[sheet].calculate_dimension()
                self.audit_file.sheets[sheet].auto_filter.enable = True
                for row in self.audit_file.sheets[sheet].iter_rows():
                    for cell in row:
                        cell.border = thin_border
        self.audit_file.close()
        # # Complete Report Sheet
        # df_style = df_temp.style.apply(lambda x: highlight_even_row(row=x))
        # df_style = df_style.apply(lambda x: gs_audit_report_color_rows(row=x), axis=1)
        # df_style.to_excel(excel_writer=self.audit_file, sheet_name='Report', index=False)
        # self.audit_file.sheets['Report'].auto_filter.ref = self.audit_file.sheets['Report'].calculate_dimension()
        # self.audit_file.sheets['Report'].auto_filter.enable = True
        # # GS_Error
        # if gs_admin:
        #     df = df_temp.loc[(~df_temp.flag) & (df_temp.GSValue != 'GS_Error')]
        #     if len(df.index) > 0:
        #         df_style = df.style.apply(lambda x: highlight_even_row(row=x))
        #         df_style = df_style.apply(lambda x: gs_audit_report_color_rows(row=x), axis=1)
        #         df_style.to_excel(excel_writer=self.audit_file, sheet_name='Delta', index=False)
        #         self.audit_file.sheets['Delta'].auto_filter.ref = self.audit_file.sheets['Delta'].calculate_dimension()
        #         self.audit_file.sheets['Delta'].auto_filter.enable = True
        #     df = df_temp.loc[df_temp.GSValue == 'Delta']
        #     if len(df.index) > 0:
        #         df_style = df.style.apply(lambda x: highlight_even_row(row=x))
        #         df_style = df_style.apply(lambda x: gs_audit_report_color_rows(row=x), axis=1)
        #         df_style.to_excel(excel_writer=self.audit_file, sheet_name='Error', index=False)
        #         self.audit_file.sheets['Error'].auto_filter.ref = self.audit_file.sheets['Error'].calculate_dimension()
        #         self.audit_file.sheets['Error'].auto_filter.enable = True
        #
        #     para_dict = {'USID': 'para', 'SITE': 'sites', 'CELL': 'cells', 'EARFCN': 'earfcn', 'ARFCN': 'ssbfreq'}
        #     for sh in ['para', 'sites', 'cells', 'earfcn', 'ssbfreq']:
        #         temp_list = []
        #         if sh == 'para':
        #             df_temp = pd.DataFrame([self.usid.param_dict.get(sh, {})])
        #         elif sh == 'sites':
        #             df_temp = pd.DataFrame([self.usid.param_dict.get(sh).get(site).get('para')
        #                                     for site in self.usid.param_dict.get(sh, {}).keys()])
        #         elif sh == 'cells':
        #             for site in self.usid.param_dict.get('sites').keys():
        #                 for cell in self.usid.param_dict.get('sites').get(site).get(sh, {}).keys():
        #                     temp_list.append(self.usid.param_dict.get('sites').get(site).get(sh).get(cell))
        #             df_temp = pd.DataFrame(temp_list)
        #         else:
        #             for key in self.usid.param_dict.get(sh, {}).keys():
        #                 tmp = self.usid.param_dict.get(sh, {}).get(key)
        #                 tmp['freq'] = key
        #                 temp_list.append(tmp)
        #             df_temp = pd.DataFrame(temp_list)
        #         if len(df_temp.index) > 0:
        #             df_temp.columns = df_temp.columns.astype(str)
        #             df_temp = df_temp.add_prefix('_')
        #             df_temp = df_temp.style.apply(lambda x: highlight_even_row(row=x))
        #             df_temp.to_excel(excel_writer=self.audit_file, sheet_name=sh, index=False)
        #             self.audit_file.sheets[sh].auto_filter.ref = self.audit_file.sheets[sh].calculate_dimension()
        #             self.audit_file.sheets[sh].auto_filter.enable = True
        #     para_dict = {'lte_freq': self.usid.df_lte_freq, 'lte_rel': self.usid.df_lte_rel, 'lte_crel': self.usid.df_lte_crel,
        #                  # 'lte_umts_freq': self.usid.df_lte_umts_freq, 'lte_umts_rel': self.usid.df_lte_umts_rel,
        #                  'lte_nr_freq': self.usid.df_lte_nr_freq, 'lte_nr_rel': self.usid.df_lte_nr_rel, 'lte_nr_crel': self.usid.df_lte_nr_crel,
        #                  'nr_freq': self.usid.df_nr_freq, 'nr_rel': self.usid.df_nr_rel, 'nr_crel': self.usid.df_nr_crel}
        #     for sheet in para_dict.keys():
        #         df_temp = para_dict[sheet].copy(deep=True)
        #         if len(df_temp.index) > 0:
        #             df_temp.columns = df_temp.columns.astype(str)
        #             df_temp = df_temp.style.apply(lambda x: highlight_even_row(row=x))
        #             df_temp.to_excel(excel_writer=self.audit_file, sheet_name=sheet, index=False)
        #             self.audit_file.sheets[sheet].auto_filter.ref = self.audit_file.sheets[sheet].calculate_dimension()
        #             self.audit_file.sheets[sheet].auto_filter.enable = True
        # self.audit_file.close()


    @staticmethod
    def excell_table_formating(work_sheet, sheet_name):
        tab = Table(displayName=sheet_name + "_Table", name=sheet_name + "_Table", ref=work_sheet.calculate_dimension())
        style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=True, showLastColumn=True, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        work_sheet.add_table(tab)

    # @staticmethod
    # def report_create_formating(work_sheet, sheet_name):
    #     """
    #         This method fetches the worksheet dimensions and returns the dimensions plus a dictionary that contains a
    #         mapping of the sheet column names to their cell coordinates
    #     """
    #     dims = work_sheet.calculate_dimension()
    #     matched = re.match(r"([A-Z]+)[\d]+:([A-Z]+)[\d]+", dims)
    #     column_ref = ":".join([str_ + '1' for str_ in matched.groups()])
    #     col_coords = dict([(cell_.value, cell_.coordinate) for cell_ in work_sheet[column_ref][0]])
    #
    #     # Add Red Colour for error flag
    #     dxf_error = DifferentialStyle(fill=PatternFill("solid", bgColor="FF0000"), font=Font(b=False, color="FFFFFF"))
    #     r_error = Rule(type="expression", dxf=dxf_error, stopIfTrue=True)
    #     r_error.formula = [F"${col_coords['GSValue']}=\"error\""]
    #     work_sheet.conditional_formatting.add(dims, r_error)
    #
    #     # If 'flag' is False & Permission is Global or GlobalOptional, then it should be filled Red
    #     dxf_global = DifferentialStyle(fill=PatternFill("solid", bgColor="F47174"))
    #     r_global = Rule(type="expression", dxf=dxf_global, stopIfTrue=True)
    #     r_global.formula = [F"AND(OR(${col_coords['Permission']}=\"Global\", ${col_coords['Permission']}=\"GlobalOptional\", $"
    #                         F"{col_coords['Permission']}=\"GlobalOptional \"),${col_coords['flag']}=FALSE)"]
    #     work_sheet.conditional_formatting.add(dims, r_global)
    #
    #     # If 'Flag' is False & Permission is Not Auditable, then it should be filled Grey
    #     dxf_not_auditable = DifferentialStyle(fill=PatternFill("solid", bgColor="B7825F"))
    #     r_not_auditable = Rule(type="expression", dxf=dxf_not_auditable, stopIfTrue=True)
    #     r_not_auditable.formula = [F"AND(${col_coords['Permission']}=\"Not Auditable\",${col_coords['flag']}=FALSE)"]
    #     work_sheet.conditional_formatting.add(dims, r_not_auditable)
    #
    #     # If 'Flag' is False or True & Permission is Local, then it should be filled Light Blue
    #     dxf_local = DifferentialStyle(fill=PatternFill("solid", bgColor="93CAED"))
    #     r_local = Rule(type="expression", dxf=dxf_local, stopIfTrue=True)
    #     r_local.formula = [F"AND(${col_coords['Permission']}=\"Local\",${col_coords['flag']}=FALSE)"]
    #     work_sheet.conditional_formatting.add(dims, r_local)
    #
    #     # If 'Flag' is False or True & Permission is Internal, then it should be filled Orange
    #     dxf_internal = DifferentialStyle(fill=PatternFill("solid", bgColor="F5CA7B"))
    #     r_internal = Rule(type="expression", dxf=dxf_internal, stopIfTrue=True)
    #     r_internal.formula = [F"AND(${col_coords['Permission']}=\"Internal\",${col_coords['flag']}=FALSE)"]
    #     work_sheet.conditional_formatting.add(dims, r_internal)
    #
    #     """
    #     This method adds a table to the excel sheet. There should be no null column names to prevent when the table is created.
    #     This is to prevent the excel sheet from getting corrupted.
    #     """
    #     tab = Table(displayName=sheet_name + "_Table", name=sheet_name + "_Table", ref=dims)
    #     # TableStyleLight15
    #     style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=True, showLastColumn=True, showRowStripes=True, showColumnStripes=False)
    #     tab.tableStyleInfo = style
    #     work_sheet.add_table(tab)
